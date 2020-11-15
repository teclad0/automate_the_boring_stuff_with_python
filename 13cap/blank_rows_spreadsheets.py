#! python3
# blank_rows_spreadsheets.py - counts the number of empty rows and empty cells
import openpyxl
from openpyxl.utils import get_column_letter

# Opens the spreadsheet
wb = openpyxl.load_workbook('SampleData.xlsx')
sheet= wb.active

# Selecting the list of tuples of rows
rows=sheet['A2':get_column_letter(sheet.max_column)+str(sheet.max_row)]
sum_rows=0
sum_cells=0

# Goes trhu all rows and checks empty ones 
for A,B,C,D,E,F,G in rows:
    if ((A.value or B.value or C.value or D.value or E.value or F.value or G.value) == None):
        sum_rows=sum_rows+1
    if ((A.value and B.value and C.value and D.value and E.value and F.value and G.value) == None):
        sum_cells=sum_cells+1

print('Empty rows: ' + str(sum_rows) +' and empty cells: '+ str(sum_cells-sum_rows))


